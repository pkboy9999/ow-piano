import sys
from itertools import chain
import xmltodict
import pyperclip
import openpyxl

A_Z = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

global beats
global beat_type

steps = {
    "C": 1,
    "D": 2,
    "E": 3,
    "F": 4,
    "G": 5,
    "A": 6,
    "B": 7
}
beat_types_map = {
    1: "whole", 2: "half", 4: "quarter", 8: "eighth", 16: "16th", 32: "32nd"
}
types = {
    "whole": 16,
    "half": 8,
    "quarter": 4,
    "eighth": 2,
    "16th": 0,
    "32nd": 0
}


def value2key(v, d):
    return list(d.keys())[list(d.values()).index(v)]


def get_note_detail(n):
    return {
        "tie": n.get("tie", {}).get("@type"),
        "staff": n.get("staff"),
        "dot": n.get("dot"),
        "rest": n.get("rest"),
        "type": n.get("type"),
        "step": n.get("pitch", {}).get("step"),
        "alter": n.get("pitch", {}).get("alter"),
        "octave": n.get("pitch", {}).get("octave"),
    }


def get_notes(m):
    _ = m.get("attributes", {}).get("time")
    # 只有第一个attributes会包含time属性
    if _ is not None:
        # 此处beats和beat_type都是外部变量
        global beats
        global beat_type
        beats = int(_.get("beats"))
        beat_type = int(_.get("beat-type"))

    if type(m["note"]) is list:
        return list(map(get_note_detail, m["note"]))
    else:
        return get_note_detail(m["note"])


def note_to_array(note):
    global beats
    global beat_type
    _ = [{
             "1": "A",
             "2": "B",
             None: "A"
         }[note["staff"]]]
    if note["tie"] == 'stop' or note["rest"] == '1':
        _.append(0)
    else:
        _.append(int(note["octave"]) * 100)
        if note["alter"] is None:
            _[-1] += steps[note["step"]]
        else:
            if note["step"] in "CDE":
                alter_note = (steps[note["step"]] + (7 if note["alter"] == '1' else 6))
            else:
                alter_note = (steps[note["step"]] + (6 if note["alter"] == '1' else 5))

            _[-1] += alter_note

    # 当rest存在时,而且type不存在的特殊情况
    if note["rest"] == "1" and note["type"] is None:
        note["type"] = value2key(beats * types[beat_types_map[beat_type]], types)
        note["type"] = 1 if note["type"] == 0 else note["type"]

    if note["dot"] is None:
        _ += (types[note["type"]] - 1) * [0]
    else:
        _ += int((types[note["type"]] - 1) + (types[note["type"]] / 2)) * [0]
    return _


if __name__ == '__main__':

    with open(sys.argv[1], encoding="utf8") as f:  # 此处取消注释可用于命令行第一个参数
    # with open("The_Truth_That_You_Leave2.musicxml", encoding="utf8") as f:  # 此处调试使用
        xml_content = f.read() \
            .replace("<dot/>", "<dot>1</dot>") \
            .replace("<rest/>", "<rest>1</rest>")
        # 由于xmltodict不会解析自闭合标签为json节点,所以dot
        # 和rest 手动替换成非自闭合标签,并且添加一个值在其中

    d = xmltodict.parse(xml_content)  # 解析xml到json
    notes = []
    # 取出measure
    measure = d.get("score-partwise", {}).get("part", {}).get("measure")

    # measure可能是单独的,也可能是一个数组
    if type(measure) is list:

        notes += list(filter(lambda x: x is not None, map(get_notes, measure)))
    else:
        notes += get_notes(measure)

    # 将notes中结果为None的剔除,并且降为一维数组                                     # notes可能是单独的,也可能是数组
    notes = list(filter(lambda x: x is not None, list(chain.from_iterable(notes)) if type(notes[0]) is list else notes))
    array = list(map(note_to_array, notes))
    print(array)
    slot1_array = list(chain.from_iterable(filter(lambda x: x[0] == "A", array)))
    slot2_array = list(chain.from_iterable(filter(lambda x: x[0] == "B", array)))

    slot1_array = list(filter(lambda x: x != "A", slot1_array))
    slot2_array = list(filter(lambda x: x != "B", slot2_array))

    print(len(slot1_array), len(slot2_array))

    def writeExcel(path, dataA, dataB):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'ow-piano'

        for i in range(0, len(dataA)):
            sheet.cell(row=i+1, column=1, value=i+1)
            sheet.cell(row=i+1, column=2, value=str(dataA[i]))
            sheet.cell(row=i+1, column=3, value=str(dataB[i]))

        wb.save(path)
        print("写入数据成功！")

    writeExcel('piano.xlsx', slot1_array, slot2_array)

    # slot1_array = [slot1_array[i:i + 1000] for i in range(0, len(slot1_array), 1000)]
    # slot2_array = [slot2_array[i:i + 1000] for i in range(0, len(slot2_array), 1000)]


    # def array_to_script(s, a):
    #     action_example = r"""
    #     Set Player Variable(Players In Slot({slot}, Team 1), {var_name},{append}Empty Array,{values});
    #     """
    #     result_text = ""
    #     for i in range(len(a)):
    #         values = ""
    #         counter = 0
    #         for member in a[i]:
    #             values += str(member) + "),"
    #             counter += 1
    #         values = values[:-1]
    #         result_text += action_example.format(
    #             slot=s,
    #             var_name=A_Z[i],
    #             append=counter * "Append To Array(",
    #             values=values
    #         )

    #     return result_text


    # text = array_to_script(0, slot1_array) + array_to_script(1, slot2_array)
    # text = "actions{" + text + "}"
    # print(text)
    # pyperclip.copy(text)
